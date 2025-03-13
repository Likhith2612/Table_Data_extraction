import os
import openpyxl
import pandas as pd
import google.generativeai as genai
import csv
import time
from openpyxl.styles import Font, Border, Side

# Configure Gemini API
API_KEY = "AIzaSyCGakY_-40t749xeqYqbc6T7ErqJkhUQbE"
genai.configure(api_key=API_KEY)
MODEL = 'gemini-1.5-pro-latest'

# --- Helper Functions ---
def is_bold(cell):
    """Check if cell font is bold"""
    return cell.font.b if cell.font else False

def has_dark_border(cell):
    """Check for dark borders"""
    dark_side = Side(border_style="thick", color="000000")
    return any([cell.border.top == dark_side,
                cell.border.bottom == dark_side,
                cell.border.left == dark_side,
                cell.border.right == dark_side])

def clean_columns(sheet, start_col, end_col):
    """Remove empty columns from processing"""
    return [col for col in range(start_col, end_col+1) 
            if not is_empty_column(sheet, col)]

# --- Core Processing Functions ---
def process_type1_table(sheet, start_row, end_row, file_path, all_data):
    """Handle bold main + normal sub-question tables"""
    headers = [cell.value for cell in sheet[start_row]]
    question_col = next((i for i, h in enumerate(headers) if h and 'question' in h.lower()), None)
    
    if question_col is None:
        return

    current_main = None
    for row_idx in range(start_row+1, end_row+1):
        cell = sheet.cell(row=row_idx, column=question_col+1)
        if is_bold(cell):
            current_main = str(cell.value).strip()
        elif current_main:
            sub_question = str(cell.value).strip() if cell.value else ""
            combined = f"{current_main} {sub_question}".strip()
            all_data.append({'question': combined, 'file': file_path})
            current_main = None

def process_type2_table(sheet, start_col, end_col, file_path, all_data):
    """Handle supporting comments as sub-questions"""
    headers = [sheet.cell(row=1, column=col).value for col in range(start_col, end_col+1)]
    question_col = next((i for i, h in enumerate(headers) if h and 'question' in h.lower()), None)
    support_col = next((i for i, h in enumerate(headers) if h and 'support' in h.lower()), None)
    
    for row_idx in range(2, sheet.max_row+1):
        main = sheet.cell(row=row_idx, column=start_col+question_col).value if question_col else ""
        support = sheet.cell(row=row_idx, column=start_col+support_col).value if support_col else ""
        
        if support and main:
            combined = f"{main.strip()} {support.strip()}".strip()
            all_data.append({'question': combined, 'file': file_path})
        elif main:
            all_data.append({'question': main.strip(), 'file': file_path})

def process_type3_table(sheet, table_range, file_path, all_data):
    """Handle control-type separated questions"""
    headers = [cell.value for cell in sheet[table_range[0]]]
    control_col = next((i for i, h in enumerate(headers) if h and 'control' in h.lower()), None)
    
    current_control = ""
    for row in sheet.iter_rows(min_row=table_range[0]+1, max_row=table_range[1]):
        if control_col and row[control_col].value:
            current_control = str(row[control_col].value).strip() + ": "
        
        question_cells = [cell for i, cell in enumerate(row) 
                         if 'question' in str(headers[i]).lower()]
        
        for cell in question_cells:
            if cell.value:
                q = f"{current_control}{cell.value.strip()}"
                all_data.append({'question': q, 'file': file_path})

# --- Main Extraction Function ---
def extract_tables(file_path):
    workbook = openpyxl.load_workbook(file_path)
    all_data = []
    
    for sheet in workbook.worksheets:
        # Detect table boundaries using blue headers
        tables = []
        for row in sheet.iter_rows():
            if row[0].fill.start_color.index == "000000FF":  # Blue header
                tables.append({
                    'start_row': row[0].row,
                    'end_row': None,
                    'type': None
                })
        
        # Process each table
        for table in tables:
            header_row = sheet[table['start_row']]
            if any('support' in str(cell.value).lower() for cell in header_row):
                process_type2_table(sheet, table['start_col'], table['end_col'], file_path, all_data)
            elif any(cell.font.b for cell in header_row):
                process_type1_table(sheet, table['start_row'], table['end_row'], file_path, all_data)
            else:
                process_type3_table(sheet, (table['start_row'], table['end_row']), file_path, all_data)
    
    return all_data

# --- Gemini Integration ---
def refine_with_gemini(questions):
    prompt = f"""Process these questions:
    {questions}
    
    1. Combine main/sub-questions into single items
    2. Remove duplicates and formatting artifacts
    3. Maintain original meaning
    4. Output as numbered list"""
    
    model = genai.GenerativeModel(MODEL)
    response = model.generate_content(prompt)
    return [q.split('. ', 1)[1] for q in response.text.split('\n') if q]

# --- Execution Flow ---
folder_path = r"C:\Users\Trellis-0039\Desktop\table_extraction\excel"
excel_files = [os.path.join(folder_path, file) for file in os.listdir(folder_path) if file.endswith('.xlsx')]

final_output = []
for file in excel_files:
    raw_data = extract_tables(file)
    refined = refine_with_gemini(raw_data)
    final_output.extend([(file, q) for q in refined])
    time.sleep(15)  # Rate limiting

# Export to CSV
pd.DataFrame(final_output, columns=['file_name', 'question'])\
  .reset_index().rename(columns={'index':'sl_no'})\
  .to_csv('output_questions.csv', index=False)
